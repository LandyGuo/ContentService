#coding=utf8
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from datetime import date
from monitor_conf import TABLE_SAVE_PATH, LOG_DATE
from MonitorDataModel import MonitorDataModel, ErrorSaveModel, ErrorInfoSaveModel, StateInfoModel

TODAY = date.today().isoformat() if LOG_DATE is 'today' else LOG_DATE

def get_data(today=TODAY):    
    return MonitorDataModel().find_table(today)

def get_error_info(threshold):
    return ErrorInfoSaveModel().get_data(TODAY, threshold)

def get_error_data(today=TODAY):
    return ErrorSaveModel().find(today)

def get_error_num(error_source,error_hour,today=TODAY):
    return ErrorSaveModel().get_num(today,error_source,error_hour)

def get_error_detail(query_type,today=TODAY):
    return ErrorSaveModel().get_error_element(today, query_type)

def get_state_info(log_type,source,hour,today=TODAY):
    return StateInfoModel().get_num(log_type, source, hour, today)

def get_state_list(query_info,today=TODAY):
    return StateInfoModel().get_state_distinct(query_info,querydate=today)
    
def dictvalue2list(dct):
    return [dct['err_num'],dct['err_type'],dct['err_info'],dct['err_location']]

def RowWriter(worksheet,start_row_pos,start_col_pos,write_list):
    row_len = len(write_list)
    ROW = start_row_pos
    for col in range(start_col_pos,start_col_pos+row_len):
        worksheet.cell(row=ROW,column=col).value = write_list[col-start_col_pos]
     
def ColWriter(worksheet,start_row_pos,start_col_pos,write_list):
    col_len = len(write_list)
    COL = start_col_pos
    for row in range(start_row_pos,start_row_pos+col_len):
        worksheet.cell(row=row,column=COL).value = write_list[row-start_row_pos]

def mktables(): 
    #新建一个工作薄
    wb = Workbook()
    #在工作薄里新建一个表单
    ws1 = wb.worksheets[0]
    #给表单取名
    ws1.title = u'资源统计'
    #从数据库中拿数据
    data = get_data()
    source_list = data['source_list']
    channel_list = data['channel_list']
    #开始表格的制作过程
    #写标题
    ws1.cell(row=0,column=0).value = u'资源统计'
    RowWriter(ws1,0,1,channel_list+[u'总计'])
    ColWriter(ws1,1,0,source_list+[u'总计'])
    #按资源的来源网站取各个channel的资源量
    current_dict = data['source_detail']
    for pos1, source in enumerate(source_list):
        for pos2,channel in enumerate(channel_list):
            num = current_dict[source][channel]['number']
            ws1.cell(row=pos1+1,column=pos2+1).value = num
    #计算channel总计
    current_dict = data['channel_info']
    write_list = []
    for channel in channel_list:
        write_list+=[current_dict[channel]['number']]
    RowWriter(ws1,len(source_list)+1,1,write_list)
    #计算source总计
    current_dict = data['source_info']
    write_list = []
    for source in source_list:
        write_list+=[current_dict[source]['number']]
    ColWriter(ws1,1,len(channel_list)+1,write_list)
    ws1.cell(row=len(source_list)+1,column=len(channel_list)+1).value = data['total_info']['number']
    ws1.cell(row=len(source_list)+2,column=len(channel_list)).value = u'统计时间'
    ws1.cell(row=len(source_list)+2,column=len(channel_list)+1).value = TODAY
    #错误分时统计
    ws2 = wb.create_sheet()
    ws2.title = u'错误统计'
    #获得所有统计数据
    source_list = get_error_detail('source')
    hour_list = get_error_detail('hour')
    #写标签
    ws2.cell(row=0,column=0).value = u'错误分时统计'
    RowWriter(ws2,0,1,source_list)
    ColWriter(ws2,1,0,hour_list)
    for pos1,source in enumerate(source_list):
        for pos2,hour in enumerate(hour_list):
            num = get_error_num(source,hour)
            ws2.cell(row=pos2+1,column=pos1+1).value = num
    #标注统计日期
    ws2.cell(row=len(hour_list)+1,column=len(source_list)-1).value = u'统计日期'
    ws2.cell(row=len(hour_list)+1,column=len(source_list)).value = TODAY
    #the end 
    #错误信息的收集与统计
    ws3 = wb.create_sheet()
    ws3.title = u'错误详情'
    #写表头
    write_list = [u'错误重复出现数',u'错误类型',u'错误信息',u'错误位置']
    RowWriter(ws3,0,0,write_list)
    #从错误信息数据库取数据
    data = get_error_info(50)
    for pos, dct in enumerate(data):
        write_list = dictvalue2list(dct)
        RowWriter(ws3,pos+1,0,write_list)
    #标注统计日期
    ws3.cell(row=pos+2,column=0).value = u"统计日期"
    ws3.cell(row=pos+2,column=1).value = TODAY
    #end
    #资源的更新数据
    ws4 = wb.create_sheet()
    ws5 = wb.create_sheet()
    for ws in (ws4, ws5):
        if ws==ws4:
            ws.title = u'更新资源'
            ws.cell(row=0,column=0).value = u'更新资源'
            info_type = 'update'
        else:
            ws.title = u'新增资源'
            ws.cell(row=0,column=0).value = u'新增资源'
            info_type = 'insert'
        #从数据库中拿数据
        data = get_data()
        source_list = data['source_list']
        channel_list = data['channel_list']
        #开始表格的制作过程
        #写标题
        RowWriter(ws,0,1,channel_list+[u'总计'])
        ColWriter(ws,1,0,source_list+[u'总计'])
        #按资源的来源网站取各个channel的资源量
        current_dict = data['source_detail']
        for pos1, source in enumerate(source_list):
            for pos2,channel in enumerate(channel_list):
                num = current_dict[source][channel][info_type]
                ws.cell(row=pos1+1,column=pos2+1).value = num
        #计算channel总计
        current_dict = data['channel_info']
        write_list = []
        for channel in channel_list:
            write_list+=[current_dict[channel][info_type]]
        RowWriter(ws,len(source_list)+1,1,write_list)
        #计算source总计
        current_dict = data['source_info']
        write_list = []
        for source in source_list:
            write_list+=[current_dict[source][info_type]]
        ColWriter(ws,1,len(channel_list)+1,write_list)
        ws.cell(row=len(source_list)+1,column=len(channel_list)+1).value = data['total_info'][info_type] 
        ws.cell(row=len(source_list)+2,column=len(channel_list)).value =u'统计时间'
        ws.cell(row=len(source_list)+2,column=len(channel_list)+1).value = TODAY
    #资源更新分时统计
    ws6 = wb.create_sheet()
    ws7 = wb.create_sheet()
    ws6.title = u'资源更新分时统计'
    ws7.title = u'资源新增分时统计'
    #获取数据库
    for ws in (ws6, ws7):
        if ws==ws6:
            log_type = 'update'
            sheet_name = u'资源更新分时统计'
        else:
            log_type = 'insert'
            sheet_name = u'资源新增分时统计'
        source_list = get_state_list('source')
        hour_list = get_state_list('hour')
        #写表头
        ws.cell(row=0,column=0).value = sheet_name
        RowWriter(ws,0,1,source_list)
        ColWriter(ws,1,0,hour_list)
        #填写每个来源每小时的更新数
        for pos1, source in enumerate(source_list):
            for pos2,hour in enumerate(hour_list):
                num = get_state_info(log_type,source,hour)
                ws.cell(row=pos2+1,column=pos1+1).value = num
        #标注统计日期
        ws.cell(row=len(hour_list)+1,column=len(source_list)-1).value = u'统计日期'
        ws.cell(row=len(hour_list)+1,column=len(source_list)).value = TODAY 
    #将所有更改一次性写入excel中
    writer = ExcelWriter(workbook=wb)
    writer.save(TABLE_SAVE_PATH)
    
    
    
if __name__=='__main__':
    mktables()